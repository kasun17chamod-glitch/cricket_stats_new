"""One-shot script to (re)generate cricket_stats.xlsx from the data in
data.py. Run after editing data.py if you want to keep the spreadsheet in
sync, or use it once to seed the workbook and then edit the spreadsheet
directly going forward."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from data import CRICKET_DATA

# Column order matches the original data.js / cricket_stats.xlsx schema.
COLUMNS = [
    "Date", "Year", "Opponent", "Match_Type",
    "Runs", "Balls", "4s", "6s",
    "Out", "Dismissal", "Catch",
    "Overs", "Runs_Conceded", "Maidens", "Wickets",
    "matchId",
]

OUT_PATH = Path(__file__).parent / "cricket_stats.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "matches"

# Header row
header_font = Font(bold=True, color="FFFFFFFF", name="Calibri")
header_fill = PatternFill("solid", fgColor="FF0A0E1A")  # match dashboard ink-900
center = Alignment(horizontal="center", vertical="center")

for c, col in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=c, value=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center

# Data rows
for r, rec in enumerate(CRICKET_DATA, start=2):
    for c, col in enumerate(COLUMNS, start=1):
        ws.cell(row=r, column=c, value=rec.get(col))

# Sensible column widths
widths = {
    "Date": 12, "Year": 7, "Opponent": 22, "Match_Type": 18,
    "Runs": 7, "Balls": 7, "4s": 5, "6s": 5,
    "Out": 6, "Dismissal": 16, "Catch": 7,
    "Overs": 8, "Runs_Conceded": 14, "Maidens": 9, "Wickets": 9,
    "matchId": 9,
}
for c, col in enumerate(COLUMNS, start=1):
    ws.column_dimensions[get_column_letter(c)].width = widths.get(col, 12)

# Freeze header row, add filters
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(OUT_PATH)
print(f"Wrote {OUT_PATH} with {len(CRICKET_DATA)} match rows.")
